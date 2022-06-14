# Importação de bibliotecas padrão
import math
import os
from datetime import datetime

# Importação de bibliotecas de terceiros
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, NamedStyle

# Importação de bibliotecas locais
import qrcodeGenerator

plt.rcParams["figure.figsize"] = (15,6) # Define a proporção dos gráficos gerados

# Define o número de linhas e colunas a serem "puladas" a partir do início da worksheet
ROW_OFFSET = 1
COLUMN_OFFSET = 1

# Define estilizações individuais sobre a fonte, bordas, preenchimento e alinhamento de células
fonte_rotulos = Font(name="Arial", bold=True)
thin = Side(border_style="thin", color="000000")
double = Side(border_style="medium", color="000000")
borda_rotulos = Border(top=double, bottom=double)
borda_dados = Border(top=thin, bottom=thin)
preenchimento_rotulos = PatternFill("solid", fgColor="B0C0C0C0")
alinhamento_centralizado = Alignment(horizontal="center", vertical="center")

# Cria e configura uma estilização "estilo_rotulos", para ser usada nos rótulos da tabela
estilo_rotulos = NamedStyle(name="estilo_rotulos")
estilo_rotulos.alignment = alinhamento_centralizado # Centraliza o conteúdo
estilo_rotulos.border = borda_rotulos
estilo_rotulos.font = fonte_rotulos
estilo_rotulos.fill = preenchimento_rotulos

# Cria e configura uma estilização "estilo_dados", para ser usada nas linhas de dados da tabela
estilo_dados = NamedStyle(name="estilo_dados")
estilo_dados.alignment = alinhamento_centralizado
estilo_dados.border = borda_dados

# Translada a origem das linhas da tabela segundo o valor definido por "ROW_OFFSET"
def row_adjust(row_value_without_offset):
    return row_value_without_offset + ROW_OFFSET

# Translada a origem das colunas da tabela segundo o valor definido por "COLUMN_OFFSET"
def col_adjust(column_value_without_offset):
    return column_value_without_offset + COLUMN_OFFSET

# Busca o primeiro valor não nulo em uma lista
def first_non_nan(list_values):
    for item in list_values:
        if math.isnan(float(item)) == False:
            return item

def inserir_tabela_resumo(wb, carteira, data_history):
    # Worksheet "Tabela"
    ###############################################################################

    # Aponta para a Worksheet ativa (criada por padrão)
    ws = wb.active
    # Acessa o atributo "title" da worksheet e o altera, renomeando a planilha
    ws.title = "Tabela"

    # Lista dos cabeçalhos da tabela
    header = ["Nome", "Código/Ticker", "Tipo", "Quantidade", "Valor unitário", "Valor Total"]

    # Lista das larguras de cada coluna da tabela
    column_widths = [40, 15, 15, 15, 25, 25]
    
    # Cria o cabeçalho de uma tabela
    ws.merge_cells(start_row=row_adjust(1), start_column=col_adjust(1), end_row=row_adjust(1), end_column=col_adjust(len(header))) # Merge as células imediatamente acima da tabela
    for i in range(1, len(header)+1):
        merged_cells = ws.cell(row=row_adjust(1), column=col_adjust(i))
        merged_cells.style = estilo_rotulos # Estiliza a célula como "estilo_rotulos"
    
    ws.cell(row=row_adjust(1), column=col_adjust(1), value="Ativos") # Define o texto da célula

    # Preenche o cabeçalho da tabela a partir da lista "header"
    for head in header:
        header_cell = ws.cell(row=row_adjust(2), column=col_adjust(header.index(head)+1), value=head)
        header_cell.style = estilo_rotulos # Estiliza a célula como "estilo_rotulos"

    # Inicializa uma variável que acumulará o valor total da carteira
    valor_total_carteira = 0

    # Itera entre os ativos das certeiras
    for row_num, data_ativo in enumerate(carteira.values(), 3):
        #Calcula o valor total referente a cada ativo e o adiciona ao dicionário
        valor_total = float(data_ativo.get("quantidade")) * float(data_ativo.get("valor_unitario"))
        data_ativo.update({"valor_total": valor_total})

        # Acumula o valor total da carteira
        valor_total_carteira += valor_total

        # Transcreve os itens do dicionario para a tabela, ajustando também o formato de exibição dos valores monetários
        ws.cell(row=row_adjust(row_num), column=col_adjust(1), value=data_ativo.get("nome"))
        ws.cell(row=row_adjust(row_num), column=col_adjust(2), value=data_ativo.get("ticker"))
        ws.cell(row=row_adjust(row_num), column=col_adjust(3), value=data_ativo.get("tipo"))
        ws.cell(row=row_adjust(row_num), column=col_adjust(4), value=data_ativo.get("quantidade"))
        ws.cell(row=row_adjust(row_num), column=col_adjust(5), value=round(data_ativo.get("valor_unitario"), 2))
        ws.cell(row=row_adjust(row_num), column=col_adjust(6), value=round(data_ativo.get("valor_total"), 2))

        # Estiliza todas as células de uma linha de dados como o "estilo_dados"
        for i in range(1, 7):
            data_cell = ws.cell(row=row_adjust(row_num), column=col_adjust(i))
            data_cell.style = estilo_dados
        
        # Define o number_format das células que contém valores monetários, para que sejam exibidos como, por exemplo, "R$ 4.059,90"
        for i in range(5, 7):
            data_cell = ws.cell(row=row_adjust(row_num), column=col_adjust(i))
            data_cell.number_format = "R$ #,##0.00"

    # Ajusta a largura das colunas da tabela segundo os valores da lista "column_widths"
    for i, column_width in enumerate(column_widths, col_adjust(1)):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Estiliza a última linha da tabela como "estilo_rotulos"
    ws.merge_cells(start_row=row_adjust(len(carteira)+3), start_column=col_adjust(1), end_row=row_adjust(len(carteira)+3), end_column=col_adjust(len(header)-1)) # Merge as células imediatamente acima da tabela
    for i in range(1, col_adjust(len(header)-1)):
        merged_cells = ws.cell(row=row_adjust(len(carteira)+3), column=col_adjust(i))
        merged_cells.style = estilo_rotulos
    
    ws.cell(row=row_adjust(len(carteira)+3), column=col_adjust(1), value="Valor total da carteira") # Define o texto da célula

    # Adiciona à tabela o valor total da carteira
    wallet_value_cell = ws.cell(row=row_adjust(len(carteira)+3), column=col_adjust(len(header)), value=round((valor_total_carteira), 2))
    # Estiliza a célula como "estilo_rotulos"
    wallet_value_cell.style = estilo_rotulos
    # Define o number_format da células que contém o valor total da carteira, para que seja exibido como, por exemplo, "R$ 4.059,90"
    wallet_value_cell.number_format = "R$ #,##0.00"
    
    # Retira a exibição das linhas de grade da planilha
    ws.sheet_view.showGridLines = False

    # Retorna o valor total da carteira
    return valor_total_carteira

def inserir_grafico_1(wb, carteira, data_history, valor_total_carteira):
    # Worksheet "Gráfico 1"
    ###############################################################################

    # Cria uma nova Worksheet chamada "Gráfico 1"
    ws = wb.create_sheet(title="Gráfico 1")

    # Declara listas que serão preenchidas a partir da leitura da carteira
    ativos = [] # Armazena o código de cada ativo
    participacao_percentual = [] # Armazena o percentual que representa a participação, em absoluto, de cada ativo no valor da carteira

    # Itera os ativos da carteira, adicionando os resultados às listas acima
    for ticker, carteira_ativo in carteira.items():
        ativos.append(ticker)
        participacao_percentual.append(carteira_ativo.get("valor_total")/valor_total_carteira*100)

    # Converte as duas listas em um pandas.DataFrame
    df_graf1 = pd.DataFrame(index=ativos, data=participacao_percentual)
    # Ordena o DataFrame em ordem decrescente de participação
    df_graf1_sorted = df_graf1.sort_values(df_graf1.columns[0], ascending=True)
    
    # Atualiza a as listas de dados, agora, já ordenados
    ativos = df_graf1_sorted.index.to_list()
    participacao_percentual = df_graf1_sorted[0].to_list()
    
    # Define as cores a serem usadas para representar ações e moedas no gráfico
    cor_acao = "#D62828"
    cor_moeda = "#003049"
    
    # Configura e realiza o plot do gráfico
    fig, ax = plt.subplots()
    fig = plt.barh(ativos, participacao_percentual, zorder=2)
    # Itera sobre os ativos, colorindo cada coluna com a cor relacionada ao seu tipo
    for i in range(0, len(ativos)):
        ativo = ativos[i]
        if(carteira.get(ativo).get("tipo") == "Ação"):
            fig[i].set_color(cor_acao)
        elif(carteira.get(ativo).get("tipo") == "Moeda"):
            fig[i].set_color(cor_moeda)
    plt.title("Composição percentual da carteira em relação ao valor absoluto", fontdict={'fontsize': 20}, pad=20) # Define o título
    plt.ylabel("Participação percentual no valor total da carteira", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo y
    plt.xlabel("Ativos", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo x
    plt.grid(color="#DCDCDC", axis="x", zorder=0) # Configura a exibição de linhas de grade
    # Gera uma legenda manualmente
    custom_lines = [Line2D([0], [0], color=cor_acao, lw=4), Line2D([0], [0], color=cor_moeda, lw=4)]
    ax.legend(custom_lines, ["Ação", "Moeda"], loc=4)

    plt.savefig("./Imagens/grafico_1.png", dpi=75, bbox_inches='tight') # Salva o gráfico gerado em "grafico_1.png"

    # Retira a exibição das linhas de grade da planilha
    ws.sheet_view.showGridLines = False

    # Abre a imagem gerada que contém o gráfico
    img = Image("./Imagens/grafico_1.png")

    # Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (2, 2) -> B2
    ws.add_image(img, ws.cell(2, 2).coordinate)

def inserir_grafico_2(wb, carteira, data_history):
    # Worksheet "Gráfico 2"
    ###############################################################################
    # Cria uma nova Worksheet chamada "Gráfico 2"
    ws = wb.create_sheet(title="Gráfico 2")

    # Cria uma "deep copy" do dataframe que contém o histórico dos ativos
    df_graf2 = data_history.copy()

    # Acessa cada coluna e processa os dados para a geração do gráfico
    for column in df_graf2:
        df_graf2[column] /= first_non_nan(df_graf2[column]) # Normaliza os valores conforme o primeiro valor não nulo
        df_graf2[column] *= 100 # Transforma a variação em dados percentuais
        df_graf2[column] -= 100 # Altera o offset, tornando o primeiro valor não nulo como 0, ou seja, como referência de comparação

    # Configura e realiza o plot do gráfico
    fig = df_graf2.plot(zorder=2) # Define os dados a serem plotados
    plt.title("Variação relativa dos ativos no último ano", fontdict={'fontsize': 20}, pad=20) # Define o título
    plt.ylabel("Variação Percentual", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo y
    plt.xlabel("Variação do tempo", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo x
    plt.grid(color="#DCDCDC", zorder=0) # Configura a exibição de linhas de grade
    fig.axhline(y=0, linestyle="dashed", color="#000000", zorder=0) # Adiciona uma linha de referência em x=0
    plt.legend(loc=(1.02, 0)) # Move a legenda do gráfico para o exterior da área de exibição
    plt.tick_params(which="both", direction="inout") # Edita os marcadores de cada eixo

    plt.savefig("./Imagens/grafico_2.png", dpi=75, bbox_inches='tight') # Salva o gráfico gerado em "grafico_1.png"

    # Retira a exibição das linhas de grade da planilha
    ws.sheet_view.showGridLines = False

    # Abre a imagem gerada que contém o gráfico
    img = Image("./Imagens/grafico_2.png")

    # Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (2, 2) -> B2
    ws.add_image(img, ws.cell(2, 2).coordinate)

def inserir_grafico_3(wb, carteira, data_history):
    # Worksheet "Gráfico 3"
    ###############################################################################

    # Cria uma nova Worksheet chamada "Gráfico 3"
    ws = wb.create_sheet(title="Gráfico 3")

    # Cria uma "deep copy" do dataframe que contém o histórico dos ativos
    df_graf3 = data_history.copy()

    # Declara listas que serão preenchidas a partir da leitura do historico de dados
    valores_total_carteira = [] # Armazena o valor total da carteira para cada timestamp
    ativos_carteira = list(carteira.values()) # Armazena a lista dos tickers dos ativos

    # Itera sorbe cada linha do DataFrame
    for row_index in range(0, df_graf3.shape[0]):
        # Define uma variável acumuladora como 0
        valor_carteira_parcial = 0

        # Itera sobre o valor de cada ação, e acumula o total investido em cada ação na variável "valor_carteira_parcial"
        for ativo in ativos_carteira:
            valor_carteira_parcial += float(ativo.get("quantidade")) * float(df_graf3[ativo.get("ticker")][row_index])

        # Após percorrer o valor de cada ativo, adiciona o valor acumulado à lista
        valores_total_carteira.append(valor_carteira_parcial)

    # Cria um novo DataFrame a partir dos timestamps e dos valores totais da carteira para cada timestamp
    df_graf3 = pd.DataFrame(index=df_graf3.index, data=valores_total_carteira)

    # Configura e realiza o plot do gráfico
    fig = df_graf3.plot(zorder=2)
    plt.title("Variação do valor total da carteira ao longo do último ano", fontdict={'fontsize': 20}, pad=20) # Define o título
    plt.ylabel("Valor em reais", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo y
    plt.xlabel("Variação do tempo", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo x
    plt.grid(color="#DCDCDC", zorder=0) # Configura a exibição de linhas de grade
    plt.tick_params(which="both", direction="inout") # Define os parâmetros dos ticks
    fig.get_legend().remove() # Remove as legendas
    plt.savefig("./Imagens/grafico_3.png", dpi=75, bbox_inches='tight') # Salva o gráfico gerado em "grafico_3.png"

    # Retira a exibição das linhas de grade da planilha
    ws.sheet_view.showGridLines = False

    # Abre a imagem gerada que contém o gráfico
    img = Image("./Imagens/grafico_3.png")

    # Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (2, 2) -> B2
    ws.add_image(img, ws.cell(2, 2).coordinate)
    ws['A1'] = '* Nos dias em que alguma ação esteja com o valor nulo, não aparecerão dados no gráfico.'

def inserir_qrcode(wb, valor_total_carteira):
    # Worksheet "QR Code"
    ###############################################################################

    # Gera um QR Code a partir do "valor_total_carteira" e retorna o caminho até o arquivo gerado
    qrcode_path = qrcodeGenerator.gerar_qrcode(valor_total_carteira)

    # Move a imagem gerada para a pasta "Imagens"
    new_qrcode_path = "./Imagens/" + qrcode_path # Define o novo caminho até o arquivo
    os.replace(qrcode_path, new_qrcode_path) # Move o arquivo

    # Cria uma nova Worksheet chamada "QR Code"
    ws = wb.create_sheet(title="QR Code")

    # Retira a exibição das linhas de grade
    ws.sheet_view.showGridLines = False

    # Imprime um texto na célula de coordenadas (2, 2) -> B2
    ws.cell(2, 2, value= "Aponte a câmera do celular para o QR Code abaixo e confira o valor total da sua carteira:")

    # Abre a imagem gerada que contém o QR code
    img = Image(new_qrcode_path)

    # Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (4, 3) -> C4
    ws.add_image(img, ws.cell(4, 3).coordinate)

# Gera um arquivo ".xlsx" contendo uma tabela, 3 gráficos e um QR code
def gerar_xlsx(carteira, data_history):
    # Cria um Workbook
    wb = Workbook()

    # Insere uma tabela contendo um resumo dos dados da carteira
    valor_total_carteira = inserir_tabela_resumo(wb, carteira, data_history)

    # Checa se a pasta "Resultados" já existe no diretório
    if not os.path.exists("Imagens"):
        # Se ainda não existia, ela é criada
        os.mkdir("Imagens")

    # Insere um gráfico representando a composição da carteira
    inserir_grafico_1(wb, carteira, data_history, valor_total_carteira)

    # Insere um gráfico representando a variação do valor dos ativos da carteira no último ano
    inserir_grafico_2(wb, carteira, data_history)

    # Insere um gráfico representando a variação do valor total da carteira no último ano
    inserir_grafico_3(wb, carteira, data_history)

    # Insere um QR code que mostra ao usuário o valor total da sua carteira
    inserir_qrcode(wb, valor_total_carteira)

    # Checa se a pasta "Resultados" já existe no diretório
    if not os.path.exists("Resultados"):
        # Se ainda não existia, ela é criada
        os.mkdir("Resultados")

    # Concatena strings para formar o nome do arquivo, o qual registra, em seu título, o momento em que os seus dados foram processados
    filename = "./Resultados/Resultado - " + str(datetime.now().strftime("%d-%m-%Y %H-%M-%S")) + ".xlsx"

    # Salva o WookBook
    wb.save(filename)

    return os.path.abspath(filename)
