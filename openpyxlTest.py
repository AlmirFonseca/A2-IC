
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

import yfinance as yf
import math
import pandas as pd
import os
from datetime import datetime

import qrcodeGenerator

import matplotlib.pyplot as plt
plt.rcParams["figure.figsize"] = (15,6)

# Translada a origem das linhas da tabela segundo o valor definido por "row_offset"
def row_adjust(row_value_without_offset):
    return row_value_without_offset + row_offset

# Translada a origem das colunas da tabela segundo o valor definido por "column_offset"
def col_adjust(column_value_without_offset):
    return column_value_without_offset + column_offset

def first_non_nan(list_values):
    for item in list_values:
        if math.isnan(float(item)) == False:
            return item

# Cria um Workbook
wb = Workbook()

# Worksheet "Dashboard"
###############################################################################

# Aponta para a Worksheet ativa
ws = wb.active
# Acessa o atributo "title" da worksheet e o altera, renomeando a planilha
ws.title = "Dashboard"

# Para uma pessoa com 22 ações de 9.23 dólares da stone, que valem 22x9.23=203,06:
# E algo semelhante em relação à amazon
header = ["Nome", "Código/Ticker", "Tipo", "Quantidade", "Valor unitário", "Valor Total"]

# Lista das larguras de cada coluna da tabela
column_widths = [40, 15, 15, 15, 25, 25]

# Carteira gerada a partir da busca através do yfinance
carteira = {
    'AMZN': {
        'ticker': 'AMZN',
        'quantidade': 10,
        'tipo': 'Ação',
        'nome': 'Amazon.com, Inc.',
        'valor_unitario': 2302.93,
        'valor_total': 23029.3
    },
    'STNE': {
        'ticker': 'STNE',
        'quantidade': 2000,
        'tipo': 'Ação',
        'nome': 'StoneCo Ltd.',
        'valor_unitario': 10.28,
        'valor_total': 205.6
    },
    'USDBRL=X': {
        'ticker': 'USDBRL=X',
        'quantidade': 1000,
        'tipo': 'Moeda',
        'nome': 'USD/BRL',
        'valor_unitario': 4.7374,
        'valor_total': 4737.4
    },
    'EURBRL=X': {
        'ticker': 'EURBRL=X',
        'quantidade': 1000,
        'tipo': 'Moeda',
        'nome': 'EUR/BRL',
        'valor_unitario': 5.0751,
        'valor_total': 5075.1
    },
    'MSFT': {
        'ticker': 'MSFT',
        'quantidade': 10,
        'tipo': 'Ação',
        'nome': 'Microsoft Corporation',
        'valor_unitario': 273.24,
        'valor_total': 2732.4
    }
}

tickers = list(carteira.keys())
tickers_str = " ".join(tickers)
ticker_num = len(carteira.keys())

if ticker_num == 0:
    print("carteira vazia")
    # TODO
    # exibir uma mensagem ou salvar um xlsx com uma frase constando uma carteira vazia
    #return 0
    
data_download = yf.download(tickers_str, period="1y", interval="1d")

data_history = pd.DataFrame(index=data_download.index)

if ticker_num == 1:
    for ticker in tickers:
        data_history[ticker] = data_download["Close"]
elif ticker_num > 1:
    for ticker in tickers:
        data_history[ticker] = data_download["Close"][ticker]
        
# Define o número de linhas e colunas a serem "puladas" a partir do início da worksheet
row_offset = 1
column_offset = 1

# Cria o cabeçalho de uma tabela
ws.merge_cells(start_row=row_adjust(1), start_column=col_adjust(1), end_row=row_adjust(1), end_column=col_adjust(len(header))) # Merge as células imediatamente acima da tabela
ws.cell(row=row_adjust(1), column=col_adjust(1), value="Ativos") # Define o texto da célula
ws.cell(row=row_adjust(1), column=col_adjust(1)).alignment = Alignment(horizontal="center") # Centraliza o conteúdo

# Preenche o cabeçalho da tabela a partir da lista "header"
for head in header:
    ws.cell(row=row_adjust(2), column=col_adjust(header.index(head)+1), value=head)

# Inicializa uma variável que acumulará o valor total da carteira
valor_total_carteira = 0

# Itera entre os ativos das certeiras
for row_num, data_ativo in enumerate(carteira.values(), 3):
    #Calcula o valor total referente a cada ativo e o adiciona ao dicionário
    valor_total = data_ativo.get("quantidade") * data_ativo.get("valor_unitario")
    data_ativo.update({"valor_total": valor_total})

    # Acumula o valor total da carteira
    valor_total_carteira += valor_total

    # Transcreve os itens do dicionario para a tabela, ajustando também o formato de exibição dos valores monetários
    ws.cell(row=row_adjust(row_num), column=col_adjust(1), value=data_ativo.get("nome"))
    ws.cell(row=row_adjust(row_num), column=col_adjust(2), value=data_ativo.get("ticker"))
    ws.cell(row=row_adjust(row_num), column=col_adjust(3), value=data_ativo.get("tipo"))
    ws.cell(row=row_adjust(row_num), column=col_adjust(4), value=data_ativo.get("quantidade"))
    ws.cell(row=row_adjust(row_num), column=col_adjust(5), value=round(data_ativo.get("valor_unitario"), 2)).number_format = "R$ #,###.00"
    ws.cell(row=row_adjust(row_num), column=col_adjust(6), value=round(data_ativo.get("valor_total"), 2)).number_format = "R$ #,###.00"

# Ajusta a largura das colunas da tabela segundo os valores da lista "column_widths"
for i, column_width in enumerate(column_widths, col_adjust(1)):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Prepara a última linha da tabela
ws.merge_cells(start_row=row_adjust(len(carteira)+3), start_column=col_adjust(1), end_row=row_adjust(len(carteira)+3), end_column=col_adjust(len(header)-1)) # Merge as células imediatamente acima da tabela
ws.cell(row=row_adjust(len(carteira)+3), column=col_adjust(1), value="Valor total da carteira") # Define o texto da célula
ws.cell(row=row_adjust(len(carteira)+3), column=col_adjust(1)).alignment = Alignment(horizontal="center") # Centraliza o conteúdo

# Imprime o valor total da carteira
ws.cell(row=row_adjust(len(carteira)+3), column=col_adjust(len(header)), value=round((valor_total_carteira), 2)).number_format = "R$ #,###.00"
        
# Worksheet "Gráfico 1"
############################################################################### 
    
# Cria uma nova Worksheet chamada "Gráfico 1"
ws = wb.create_sheet(title="Gráfico 1")

# Declara listas que serão preenchidas a partir da leitura da carteira
ativos = [] # Armazena o código de cada ativo
participacao_percentual = [] # Armazena o percentual que representa a participação, em absoluto, de cada ativo no valor da carteira

# Itera os ativos da carteira, adicionando os resultados às listas acima
for ticker, carteira_ativo in carteira.items():
    ativos.append(ticker)
    participacao_percentual.append(carteira_ativo.get("valor_total")/valor_total_carteira*100)

# Converte as duas listas em um pandas.DataFrame
df_graf1 = pd.DataFrame(index=ativos, data=participacao_percentual)
# Ordena o DataFrame em ordem decrescente de participação
df_graf1 = df_graf1.sort_values(df_graf1.columns[0], ascending=False)

# Configura e realiza o plot do gráfico
fig = df_graf1.plot(kind="bar", zorder=2) # Define os dados a serem plotados
plt.title("Composição percentual da carteira em relação ao valor absoluto", fontdict={'fontsize': 20}, pad=20) # Define o título
plt.ylabel("Participação percentual no valor total da carteira", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo y
plt.xlabel("Ativos", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo x
plt.grid(color="#DCDCDC", axis="y", zorder=0) # Configura a exibição de linhas de grade
fig.get_legend().remove() # Remove as legendas

plt.savefig("grafico_1.png", dpi=75, bbox_inches='tight') # Salva o gráfico gerado em "grafico_1.png"

# Retira a exibição das linhas de grade da planilha
ws.sheet_view.showGridLines = False

# Abre a imagem gerada que contém o QR code
img = Image("grafico_1.png")

# Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (2, 2) -> B2
ws.add_image(img, ws.cell(2, 2).coordinate)

# Worksheet "Gráfico 2"
###############################################################################     
# Cria uma nova Worksheet chamada "Gráfico 2"
ws = wb.create_sheet(title="Gráfico 2")

# Cria uma "deep copy" do dataframe que contém o histórico dos ativos
df_graf2 = data_history.copy()

# Acessa cada linha e processa os dados para a geração do gráfico
for column in df_graf2:
    df_graf2[column] /= first_non_nan(df_graf2[column]) # Normaliza os valores conforme o primeiro valor não nulo
    df_graf2[column] *= 100 # Transforma a variação em dados percentuais
    df_graf2[column] -= 100 # Altera o offset, tornando o primeiro valor não nulo como 0, ou seja, como referência de comparação

# Configura e realiza o plot do gráfico
fig = df_graf2.plot(zorder=2) # Define os dados a serem plotados
plt.title("Variação relativa dos ativos no último ano", fontdict={'fontsize': 20}, pad=20) # Define o título
plt.ylabel("Variação Percentual", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo y
plt.xlabel("Variação do tempo", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo x
plt.grid(color="#DCDCDC", zorder=0) # Configura a exibição de linhas de grade
fig.axhline(y=0, linestyle="dashed", color="#000000", zorder=0) # Adiciona uma linha de referência em x=0
plt.legend(loc=(1.02, 0)) # Move a legenda do gráfico para o exterior da área de exibição
plt.tick_params(which="both", direction="inout") # Edita os marcadores de cada eixo

plt.savefig("grafico_2.png", dpi=75, bbox_inches='tight') # Salva o gráfico gerado em "grafico_1.png"

# Retira a exibição das linhas de grade da planilha
ws.sheet_view.showGridLines = False

# Abre a imagem gerada que contém o QR code
img = Image("grafico_2.png")

# Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (2, 2) -> B2
ws.add_image(img, ws.cell(2, 2).coordinate)

# Worksheet "Gráfico 3"
###############################################################################

# Cria uma nova Worksheet chamada "Gráfico 3"
ws = wb.create_sheet(title="Gráfico 3")

# Cria uma "deep copy" do dataframe que contém o histórico dos ativos
df_graf3 = data_history.copy()

# Declara listas que serão preenchidas a partir da leitura do historico de dados
valores_total_carteira = [] # Armazena o valor total da carteira para cada timestamp
ativos_carteira = list(carteira.values()) # Armazena a lista dos tickers dos ativos

# Itera sorbe cada linha do DataFrame
for row_index in range(0, df_graf3.shape[0]):
    # Define uma variável acumuladora como 0
    valor_carteira_parcial = 0
    
    # Itera sobre o valor de cada ação, e acumula o total investido em cada ação na variável "valor_carteira_parcial"
    for ativo in ativos_carteira: 
        valor_carteira_parcial += ativo.get("quantidade") * df_graf3[ativo.get("ticker")][row_index]
        
    # Após percorrer o valor de cada ativo, adiciona o valor acumulado à lista
    valores_total_carteira.append(valor_carteira_parcial)

# Cria um novo DataFrame a partir dos timestamps e dos valores totais da carteira para cada timestamp
df_graf3 = pd.DataFrame(index=df_graf3.index, data=valores_total_carteira)

# Configura e realiza o plot do gráfico
fig = df_graf3.plot(zorder=2)
plt.title("Variação do valor total da carteira ao longo do último ano", fontdict={'fontsize': 20}, pad=20) # Define o título
plt.ylabel("Valor em reais", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo y
plt.xlabel("Variação do tempo", fontdict={'fontsize': 14}, labelpad=10) # Define o rótulo do eixo x
plt.grid(color="#DCDCDC", zorder=0) # Configura a exibição de linhas de grade
plt.tick_params(which="both", direction="inout") # Define os parâmetros dos ticks
fig.get_legend().remove() # Remove as legendas
plt.savefig("grafico_3.png", dpi=75, bbox_inches='tight') # Salva o gráfico gerado em "grafico_3.png"

# Retira a exibição das linhas de grade da planilha
ws.sheet_view.showGridLines = False

# Abre a imagem gerada que contém o QR code
img = Image("grafico_3.png")

# Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (2, 2) -> B2
ws.add_image(img, ws.cell(2, 2).coordinate)
   
# Worksheet "QR Code"
###############################################################################

# Gera um QR Code a partir do "valor_total_carteira" e retorna o caminho até o arquivo gerado
qrcode_path = qrcodeGenerator.gerar_qrcode(valor_total_carteira)

# Cria uma nova Worksheet chamada "QR Code"
ws = wb.create_sheet(title="QR Code")

# Retira a exibição das linhas de grade
ws.sheet_view.showGridLines = False

# Imprime um texto na célula de coordenadas (2, 2) -> B2
ws.cell(2, 2, value= "Aponte a câmera do celular para o QR Code abaixo e confira o valor total da sua carteira:")

# Abre a imagem gerada que contém o QR code
img = Image(qrcode_path)

# Adiciona a imagem à tabela, ancorada sobre a célula de coordenadas (4, 3) -> C4
ws.add_image(img, ws.cell(4, 3).coordinate)

# Checa se a pasta "Resultados" já existe no diretório
if not os.path.exists("Resultados"):
    # Se ainda não existia, ela é criada
    os.mkdir("Resultados")

# Concatena strings para formar o nome do arquivo, o qual registra, em seu título, o momento em que os seus dados foram processados
filename = "./Resultados/Resultado - " + str(datetime.now().strftime("%d-%m-%Y %H-%M-%S")) + ".xlsx"

# Salva o WookBook
wb.save(filename)

# TODO PROTEGER A TABELA?